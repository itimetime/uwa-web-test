import cf.config
from openpyxl import load_workbook
import os
from multiprocessing import Process
import multiprocessing

# def ReadConfig():
#     curPath = os.path.dirname(os.path.realpath(__file__))
#     configPath = os.path.join(curPath, 'cf/cf.ini')
#     conf = configparser.ConfigParser()
#     conf.read(configPath)
#     casePath = conf.get('casepath', 'CASEPATH')
#     return casePath
def get_files():
    return os.listdir(cf.config.test_case_path)



def GetTestCasePath(file_name):
    return os.path.join(cf.config.test_case_path, file_name)


def ReadTestCase(casePath):
    caseList = []

    wb = load_workbook(casePath)
    sheetlist = wb.get_sheet_names()
    for i in sheetlist:
        sheet = wb.get_sheet_by_name(i)
        sheetrowmax = sheet.max_row
        sheetcolumnmax = sheet.max_column
        for _row in range(2, sheetrowmax + 1):
            _case = []
            for _column in range(1, sheetcolumnmax + 1):
                _case.append(sheet.cell(row=_row, column=_column).value)
            _case = tuple(_case)
            # caseDict = dict(zip(caseTitle, _case))
            # caseList.append(caseDict)
            caseList.append((_case))

    return caseList


def get_case(case_name):
    case_path = GetTestCasePath(case_name)
    ret = ReadTestCase(case_path)
    return ret


def run_task(file_name):
    cmd_str = "python testrun.py " + file_name
    print(cmd_str)
    os.system(cmd_str)


if __name__ == '__main__':
    # p = Process(target=run_task, args=(files[0],))
    # p.start()
    for i in get_files():
        run_task(i)
        # p = i.split(".")[0]
        # print(p)
        # a = Process(target=run_task, args=(i,))
        # a.start()
        # a.join()
    # for i in files:
    #     i = i.split(".")[0]
    #     eval(i + ".start()")
    #     eval(i + ".join()")
    # os.system("python testrun.py GOT.xlsx")
